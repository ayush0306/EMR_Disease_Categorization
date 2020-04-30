import re
import sys
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from extract import *

class WorkbookHandler:
    def __init__(self, workbook_path, non_nullable_fields_map):
        self.non_nullable_fields_map = non_nullable_fields_map
        print('loading workbook')
        self.work_book = load_workbook(workbook_path, data_only=True, keep_links=True)
        print('workbook loaded')
    def load_keys(self, work_sheet):
        return [col[0].value for col in work_sheet.columns]
    def is_valid_row(self, sheet_name, row):
        for field in self.non_nullable_fields_map[sheet_name]:
            if field in row and (not row[field] or not row[field].strip() or row[field] == 'NULL'):
                return False
        return True
    def iterate(self, sheet_name):
        work_sheet = self.work_book[sheet_name]
        keys = self.load_keys(work_sheet)
        for row_idx, row in enumerate(work_sheet.rows):
            # Don't make json for 1st row
            if not row_idx:
                continue
            row_dict = {}
            for idx, cell in enumerate(row):
                row_dict[keys[idx]] = str(cell.value) if cell.value else cell.value
            if not self.is_valid_row(sheet_name, row_dict):
                continue
            yield row_dict, row, work_sheet
    def write(self, out_file):
        self.work_book.save(filename=out_file)

class EyeSmartWorkbookHandler(WorkbookHandler):
    # Store	ItemCategoryCode	Brand	Company	ItemCode	PackName	Description	UnitPrice	Availability
    HISTORY = 'tocs_past_history'
    DISEASE = 'Disease (Key Word)'
    KEY_WORDS = 'Key Words (other)'
    DURATION = 'Duration (since)'
    PERIOD = 'Months/ Years'
    SHEET_DIAGNOSIS_REQUIRED = "Diagnosis Required"
    SHEET_DATA = "788345"
    def __init__(self, workbook_path):
        super().__init__(workbook_path, {
            EyeSmartWorkbookHandler.SHEET_DATA: [
                EyeSmartWorkbookHandler.HISTORY
            ],
            EyeSmartWorkbookHandler.SHEET_DIAGNOSIS_REQUIRED: [
                EyeSmartWorkbookHandler.DISEASE
            ]
        })
    def extract(self):
        all_extracted = []
        diseases_and_synonyms = []
        diseases_map = {}
        for (d, row, sheet) in self.iterate(EyeSmartWorkbookHandler.SHEET_DIAGNOSIS_REQUIRED):
            disease = d[EyeSmartWorkbookHandler.DISEASE]
            keywords = []
            if d[EyeSmartWorkbookHandler.KEY_WORDS]:
                keywords.extend([k.strip() for k in d[EyeSmartWorkbookHandler.KEY_WORDS].split(",")])
            diseases_map[disease] = keywords
            diseases_and_synonyms.append(disease)
            diseases_and_synonyms.extend(keywords)
        diseases_and_synonyms = list(filter(lambda disease: disease, diseases_and_synonyms))
        list_diseases = list(diseases_map.keys())
        list_diseases = [x.lower() for x in list_diseases]
        list_diseases = list( dict.fromkeys(list_diseases) )
        for (d, row, sheet) in self.iterate(EyeSmartWorkbookHandler.SHEET_DATA):
            history = d[EyeSmartWorkbookHandler.HISTORY]
            # print(history, diseases_and_synonyms)
            extracted = extract_disease_and_time(history, diseases_and_synonyms, row)
            all_extracted.append(extracted)
            print("Output:",extracted)
            last_col_idx = column_index_from_string(row[-1].column)
            # print(last_col_idx)
            for match in extracted:
                sheet.cell(column=last_col_idx+1, row=row[-1].row, value=match["disease"])
                sheet.cell(column=last_col_idx+2, row=row[-1].row, value=match["duration"])
                sheet.cell(column=last_col_idx+3, row=row[-1].row, value=match["duration_unit"])
                last_col_idx = last_col_idx+3
        save_excel(list_diseases,all_extracted)
        return all_extracted


# def extract_disease_and_time(history, diseases):
#     diseases_re = f'({"|".join(diseases)})'
#     # TIME HANDLING
#     # In the year of - NOT HANDLED
#     # Since childhood
#     # Since 2 years
#     # Since 2 year(s)
#     # Since 1.5 Months
#     # Since 1.5 Month(s)
#     time_re = '((?:\bsince\b){0,1}[\s]+(([0-9]{1,2})[\s]+(year|month|yr|mo)[\(s\)|s]|childhood))'
#     matches = re.findall(diseases_re+".*"+time_re, history, flags=re.IGNORECASE)
#     ms = []
#     for match in matches:
#         # match is like:
#         # ('Hypertension', ' 6 Year(', '6 Year(', '6', 'Year')
#         # ('HTN', ' childhood', 'childhood', '', '')
#         ms.append({
#             "disease": match[0],
#             "duration": match[3] or match[2],
#             "duration_unit": {
#                 'yr': 'Year',
#                 'mo': 'Month',
#             }.get(match[4] or match[1], match[4] or match[1])
#         })
#     return ms


def __main__():
    if len(sys.argv) == 1:
        raise Exception("requires input filename")

    eye_smart_workbook_handler = EyeSmartWorkbookHandler(sys.argv[1])
    all_extracted = eye_smart_workbook_handler.extract()
    print(all_extracted[0:15])
    # output_file = sys.argv[2] if len(sys.argv) == 3 else "out.xlsx"
    # eye_smart_workbook_handler.write(output_file)
    # print("output written to", output_file)


__main__()
